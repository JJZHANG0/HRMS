import os
import re
import json
import time
import tempfile
import requests
from datetime import datetime
from pdfminer.high_level import extract_text
from django.conf import settings
from rest_framework import status, permissions, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Candidate, CooperationRecord, Favorite
from .serializers import CandidateSerializer, CooperationRecordSerializer, FavoriteSerializer
from django.contrib.auth.models import User
import logging

# ⚙️ 屏蔽 pdfminer 噪音日志
logging.getLogger("pdfminer").setLevel(logging.ERROR)


# ✅ 日期格式修复函数
def normalize_date(value):
    """将 AI 返回的毕业时间统一为 YYYY-MM-DD 格式"""
    if not value:
        return None
    try:
        value = str(value).strip()

        # 完整格式
        if re.match(r"^\d{4}-\d{2}-\d{2}$", value):
            return value
        # 年-月 或 年/月
        if re.match(r"^\d{4}[-/]\d{2}$", value):
            year, month = re.split(r"[-/]", value)
            return f"{year}-{month}-01"
        # 只有年份
        if re.match(r"^\d{4}$", value):
            return f"{value}-01-01"
        # 中文格式（例如 “2027年7月”）
        match = re.match(r"^(\d{4})年(\d{1,2})月", value)
        if match:
            year, month = match.groups()
            return f"{year}-{int(month):02d}-01"
    except Exception:
        return None
    return None


class ResumeUploadView(APIView):
    """
    上传简历 PDF → 提取字段 → AI 评分 → 存数据库
    """
    permission_classes = [permissions.AllowAny]

    API_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
    API_KEY = "sk-a99ede93ae2948928ea5b10133538a9b"
    HEADERS = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}",
    }

    def post(self, request, *args, **kwargs):
        files = request.FILES.getlist("files")

        if not files:
            return Response({"error": "未上传文件"}, status=status.HTTP_400_BAD_REQUEST)
        if len(files) > 10:
            return Response({"error": "一次最多上传 10 份简历"}, status=status.HTTP_400_BAD_REQUEST)

        created, failed = [], []

        for idx, file in enumerate(files, 1):
            print(f"\n📄 [{idx}/{len(files)}] 正在处理：{file.name}")
            temp_path = None

            try:
                # 临时保存 PDF
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
                    for chunk in file.chunks():
                        temp_file.write(chunk)
                    temp_path = temp_file.name

                # 提取文字
                text = extract_text(temp_path).strip()
                if len(text) < 40:
                    print(f"⚠️ {file.name} 内容过短，跳过。")
                    failed.append({"file": file.name, "reason": "内容过短"})
                    continue

                # 调用 AI
                ai_result = self.call_ai_parse(text)
                if not ai_result:
                    failed.append({"file": file.name, "reason": "AI 无返回或解析失败"})
                    continue

                result_json = ai_result.get("data", {})
                match_level = ai_result.get("score", "D")

                # 存入数据库
                candidate = Candidate.objects.create(
                    name=result_json.get("name", ""),
                    gender=result_json.get("gender", ""),
                    age=result_json.get("age"),
                    phone=result_json.get("phone", ""),
                    email=result_json.get("email", ""),
                    major=result_json.get("major", ""),
                    education=result_json.get("degree", "本科"),
                    university=result_json.get("university", ""),
                    graduation_date=normalize_date(result_json.get("graduation_date")),
                    base=result_json.get("base", "远程"),
                    experience=result_json.get("experience", ""),
                    cooperation_status="未合作",
                    match_level=match_level,
                )
                candidate.resume_file.save(file.name, file, save=True)

                created.append(CandidateSerializer(candidate).data)
                print(f"✅ 成功导入：{candidate.name}（评分 {match_level}）")

                time.sleep(1)  # 避免Qwen速率限制

            except Exception as e:
                print(f"❌ 解析异常：{e}")
                failed.append({"file": file.name, "reason": str(e)})
            finally:
                if temp_path and os.path.exists(temp_path):
                    os.remove(temp_path)

        return Response({
            "created": created,
            "failed": failed,
            "count": len(created)
        })

    # ==============
    # 🤖 AI 调用部分
    # ==============
    def call_ai_parse(self, text):
        """
        调用通义千问提取简历字段 + 匹配评分
        """
        prompt = f"""
你是一位专业的人才简历分析系统，专注于为教育行业筛选优质候选人。
请阅读以下简历文本，严格输出 JSON，字段如下：
{{
  "data": {{
    "name": "",
    "gender": "",
    "age": "",
    "phone": "",
    "email": "",
    "major": "",
    "degree": "",
    "university": "",
    "graduation_date": "",
    "base": "",
    "experience": [
      {{
        "company": "",
        "position": "",
        "start_date": "",
        "end_date": "",
        "description": ""
      }}
    ]
  }},
  "score": ""
}}

🎯 字段提取要求：

1. **专业（major）**：优先提取【生物/数学/物理/计算机/化学/商业/机械工程】等具体专业名称。
   如果无法确定具体专业，则按大类标记：理学、工学、经管、语言、文哲史、其他

2. **Base（所在地/期望工作地）**：
   - 优先识别：上海、杭州、广州、南京、宁波（这五个城市）
   - 如果候选人不在以上城市或未明确说明，则标记为"远程"
   - 注意：提取城市名称，不要省份（如"浙江杭州"应提取为"杭州"）

3. **工作经历（experience）**：
   - **优先提取教育相关经历**（教师、助教、教练、培训师、课程设计等）
   - 按时间倒序排列（最近的经历排在前面）
   - 每条经历必须包含：公司/机构名、职位、起止时间、工作描述

4. **毕业时间（graduation_date）**：
   - 必须是 YYYY-MM-DD 格式
   - 如果只有年月，则补充为 YYYY-MM-01
   - 如果只有年份，则补充为 YYYY-01-01

📊 **简历匹配度评分（score）**：
- **A类**：有教育行业经验 + 较匹配（如：曾任教师/教练，教育背景优秀）
- **B类**：无教育经验 + 项目经验丰富（如：有丰富实习/项目，潜力大）
- **C类**：有教育经验 + 匹配度一般（如：短期教育相关实习，经验较浅）
- **D类**：无教育经验 + 匹配度一般（如：应届生，经验较少）
- **E类**：匹配度低（如：信息不足或专业领域差异大）

注：教育行业经验特指在学校、培训机构、教育公司的教学、课程设计、教育产品相关工作。

❗务必返回纯 JSON，不要任何解释性文字。

以下是简历文本：
{text[:4000]}
"""

        payload = {
            "model": "qwen-plus",
            "messages": [{"role": "user", "content": prompt}],
        }

        resp = requests.post(self.API_URL, headers=self.HEADERS, data=json.dumps(payload))
        try:
            content = resp.json()["choices"][0]["message"]["content"]
            match = re.search(r"\{[\s\S]*\}", content)
            if not match:
                raise ValueError(f"AI 未返回 JSON 格式：{content[:200]}")
            json_str = match.group(0)
            return json.loads(json_str)
        except Exception as e:
            print("⚠️ AI响应异常：", e)
            print("🧾 原始响应：", resp.text[:800])
            return None


# ✅ 人才库列表接口
class CandidateListView(generics.ListAPIView):
    queryset = Candidate.objects.all().order_by('-created_at')
    serializer_class = CandidateSerializer


class CandidateUpdateView(generics.UpdateAPIView):
    """
    单个候选人更新接口（支持 PATCH）
    用于更新评分、合作状态等字段
    """
    queryset = Candidate.objects.all()
    serializer_class = CandidateSerializer
    permission_classes = [permissions.AllowAny]


# ✅ 统计数据接口
class CandidateStatsView(APIView):
    """
    返回候选人统计数据
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        from django.db.models import Count

        total = Candidate.objects.count()
        
        # 评分分布
        score_distribution = {
            'A': Candidate.objects.filter(match_level='A').count(),
            'B': Candidate.objects.filter(match_level='B').count(),
            'C': Candidate.objects.filter(match_level='C').count(),
            'D': Candidate.objects.filter(match_level='D').count(),
            'E': Candidate.objects.filter(match_level='E').count(),
        }
        
        # 合作状态分布
        cooperation_distribution = {}
        coop_stats = Candidate.objects.values('cooperation_status').annotate(count=Count('id'))
        for item in coop_stats:
            cooperation_distribution[item['cooperation_status']] = item['count']
        
        # Base分布
        base_distribution = {}
        base_stats = Candidate.objects.values('base').annotate(count=Count('id'))
        for item in base_stats:
            base_distribution[item['base']] = item['count']
        
        # 专业分布
        major_distribution = {}
        major_stats = Candidate.objects.values('major').annotate(count=Count('id'))
        for item in major_stats:
            major = item['major'] or '未知'
            major_distribution[major] = item['count']

        return Response({
            'total': total,
            'scoreDistribution': score_distribution,
            'cooperationDistribution': cooperation_distribution,
            'baseDistribution': base_distribution,
            'majorDistribution': major_distribution,
        })


# ✅ 导出Excel接口
class CandidateExportView(APIView):
    """
    导出候选人数据为Excel
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "候选人数据"
            
            # 标题行
            headers = [
                '姓名', '性别', '年龄', '电话', '邮箱', 
                '学历', '专业', '毕业院校', '毕业时间',
                'Base', '合作状态', '匹配度', '创建时间'
            ]
            
            # 设置标题样式
            header_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数据行
            candidates = Candidate.objects.all().order_by('-created_at')
            for row_num, candidate in enumerate(candidates, 2):
                ws.cell(row=row_num, column=1, value=candidate.name)
                ws.cell(row=row_num, column=2, value=candidate.gender or '')
                ws.cell(row=row_num, column=3, value=candidate.age)
                ws.cell(row=row_num, column=4, value=candidate.phone)
                ws.cell(row=row_num, column=5, value=candidate.email)
                ws.cell(row=row_num, column=6, value=candidate.education)
                ws.cell(row=row_num, column=7, value=candidate.major or '')
                ws.cell(row=row_num, column=8, value=candidate.university or '')
                ws.cell(row=row_num, column=9, value=str(candidate.graduation_date) if candidate.graduation_date else '')
                ws.cell(row=row_num, column=10, value=candidate.base)
                ws.cell(row=row_num, column=11, value=candidate.cooperation_status)
                ws.cell(row=row_num, column=12, value=candidate.match_level)
                ws.cell(row=row_num, column=13, value=candidate.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 准备响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=candidates_{datetime.now().strftime("%Y%m%d")}.xlsx'
            wb.save(response)
            
            return response
            
        except ImportError:
            return Response(
                {'error': '请安装 openpyxl: pip install openpyxl'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ✅ 合作记录列表与创建接口
class CooperationRecordListCreateView(APIView):
    """
    获取所有合作记录 (GET) 或创建新的合作记录 (POST)
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        """获取所有合作记录，按开始时间倒序"""
        records = CooperationRecord.objects.all().order_by('-start_date')
        serializer = CooperationRecordSerializer(records, many=True)
        return Response(serializer.data)

    def post(self, request):
        """创建新的合作记录"""
        try:
            candidate_id = request.data.get('candidate_id')
            
            if not candidate_id:
                return Response(
                    {'error': '缺少候选人ID'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # 创建合作记录
            record = CooperationRecord.objects.create(
                candidate_id=candidate_id,
                project_name=request.data.get('project_name'),
                start_date=request.data.get('start_date'),
                end_date=request.data.get('end_date') or None,
                role=request.data.get('role', ''),
                salary=request.data.get('salary', ''),
                evaluation=request.data.get('evaluation', ''),
                cooperation_result=request.data.get('cooperation_result', '良好'),
            )
            
            # 处理协议文件
            if 'agreement_file' in request.FILES:
                record.agreement_file = request.FILES['agreement_file']
                record.save()
            
            serializer = CooperationRecordSerializer(record)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ✅ 合作记录详情接口（查看/更新/删除）
class CooperationRecordDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    获取、更新或删除单个合作记录
    """
    queryset = CooperationRecord.objects.all()
    serializer_class = CooperationRecordSerializer
    permission_classes = [permissions.AllowAny]


# ✅ 收藏/取消收藏接口
class ToggleFavoriteView(APIView):
    """
    收藏或取消收藏候选人
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username')
        candidate_id = request.data.get('candidate_id')

        if not username or not candidate_id:
            return Response(
                {'error': '缺少用户名或候选人ID'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # 获取用户和候选人
            user = User.objects.get(username=username)
            candidate = Candidate.objects.get(id=candidate_id)

            # 检查是否已收藏
            favorite = Favorite.objects.filter(user=user, candidate=candidate).first()

            if favorite:
                # 已收藏，则取消收藏
                favorite.delete()
                return Response({
                    'message': '已取消收藏',
                    'is_favorited': False
                })
            else:
                # 未收藏，则添加收藏
                Favorite.objects.create(user=user, candidate=candidate)
                return Response({
                    'message': '已添加收藏',
                    'is_favorited': True
                }, status=status.HTTP_201_CREATED)

        except User.DoesNotExist:
            return Response(
                {'error': '用户不存在'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Candidate.DoesNotExist:
            return Response(
                {'error': '候选人不存在'},
                status=status.HTTP_404_NOT_FOUND
            )


# ✅ 获取我的收藏列表
class MyFavoritesView(APIView):
    """
    获取当前用户的所有收藏
    """
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        username = request.GET.get('username')

        if not username:
            return Response(
                {'error': '缺少用户名'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(username=username)
            favorites = Favorite.objects.filter(user=user).select_related('candidate')
            
            # 返回候选人数据
            candidates_data = []
            for fav in favorites:
                candidate = fav.candidate
                serializer = CandidateSerializer(candidate, context={'request': request})
                data = serializer.data
                data['favorited_at'] = fav.created_at
                data['is_favorited'] = True
                candidates_data.append(data)
            
            return Response(candidates_data)

        except User.DoesNotExist:
            return Response(
                {'error': '用户不存在'},
                status=status.HTTP_404_NOT_FOUND
            )


